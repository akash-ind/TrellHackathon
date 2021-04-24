from django.shortcuts import render, redirect
from trell.models import User, Trail, Tags, UserTagScore
from django.contrib.auth import authenticate, login as user_login
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from rest_framework.response import Response
from rest_framework.decorators import api_view
# Create your views here.


def home(request):
    return render(request, "trell/landing.html")

def login(request):
    if request.method == "POST":
        username = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(username = username, password = password)
        if user:
            user_login(request, user)
            messages.success(request, "Login Successful!!")
            return redirect('trell:profile')
        else:
            messages.warning(request, "Can't  find user")
            return redirect('trell:login')
    else:
        return render(request, 'trell/register.html')


def register(request):
    if request.method == "GET":
        return render(request, 'trell/register.html')
    else:
        posted_data = request.POST
        email = posted_data.get('email')
        if User.objects.filter(email = email).exists():
            messages.warning(request, "The user with this email already exists.Please login.")
        first_name = posted_data.get('first_name')
        last_name = posted_data.get('last_name')
        user = User(first_name = first_name, 
        last_name= last_name, email = email)
        password = posted_data.get('password')
        user.set_password(password)
        user.save()
        user_login(request, user)
        messages.success(request, "User signed up successfully")
        return redirect("trell:login")


@login_required
def profile(request):
    trails = Trail.objects.filter(user = request.user).order_by('-created')[:2]
    context = {}
    context['trails'] = trails
    print(context)
    return render(request, 'trell/profile.html', context)


@login_required
def dashboard(request):
    context = {}
    fashion_tag = Tags.objects.get(tag__iexact = 'fashion')
    fashion_trails =  Trail.objects.filter(tags= fashion_tag).order_by('-popularity')[:2]

    travel_tag = Tags.objects.get(tag__iexact = 'travel')
    travel_trails =  Trail.objects.filter(tags= travel_tag).order_by('-popularity')[:2]

    entertainment_tag = Tags.objects.get(tag__iexact = 'entertainment')
    entertainment_trails =  Trail.objects.filter(tags= entertainment_tag).order_by('-popularity')[:2]

    context['fashion_trails'] = fashion_trails
    context['travel_trails'] = travel_trails
    context['entertainment_trails'] = entertainment_trails
    return render(request, 'trell/dashboard.html', context)

@login_required
def discover(request):
    context = {}
    tag = None
    if request.GET.get('tag'):
        tag = request.GET.get('tag')
        try:
            tag = Tags.objects.get(tag__iexact=tag)
        except:
            tag = None
    if tag:
        trails = Trail.objects.filter(tags = tag).exclude(watched_by = request.user).order_by('-popularity')[:20]
    else:
        trails = Trail.objects.exclude(watched_by = request.user).order_by('-popularity')[:20]
    context['trails'] = trails
    tags = Tags.objects.all()
    context['tags'] = tags
    return render(request, 'trell/discover.html', context)


@login_required
@api_view(["POST"])
def add_to_watched(request):
    user = request.user
    trail_id = request.data.get('id')
    print(trail_id)
    trail = Trail.objects.get(trail_id = trail_id)
    trail.watched_by.add(user)
    return Response({'watched':True})



from openpyxl import load_workbook
from django.conf import settings
import os


@login_required
def populate_trails(request):
    wb = load_workbook(os.path.join(settings.BASE_DIR, 'Trell.xlsx'))
    sheet = wb['Sheet1']
    for i in range(2, 103):
        id = int(float(sheet[f'A{i}'].value))
        trail = Trail(trail_id = id)
        trail.user =  request.user
        trail.save()

    return redirect('trell:dashboard')

@login_required
def load_users(request):
    wb = load_workbook(os.path.join(settings.BASE_DIR, 'Trell_Final.xlsx'))
    sheet = wb['Users']
    tags = Tags.objects.all()
    for i in range(2, 455):
        user_id = int(sheet[f'A{i}'].value)
        user = User(user_id = user_id)
        user.email = f'some{i}@email.com'
        user.set_password('akash')
        user.save()
        for j in range(13):
            letter = chr(ord('B')+j)
            tag_score = UserTagScore(user = user)
            tag_score.score = int(sheet[f'{letter}{i}'].value)
            tag_score.tag = tags[j]
            tag_score.save()
    return redirect("trell:dashboard")

@login_required
def load_trails(request):
    wb = load_workbook(os.path.join(settings.BASE_DIR, 'Trell_Final-1.xlsx'))
    sheet = wb['Sheet2']
    tags = Tags.objects.all()
    for i in range(2, 305):
        trail_id = int(sheet[f'A{i}'].value)
        trail = Trail(trail_id = trail_id)
        trail.title =  f'trail-{i}'
        trail.user = request.user
        trail.popularity = float(sheet[f'O{i}'].value)
        try:
            trail.save()
        except:
            pass
        for j in range(13):
            letter = chr(ord('B')+j)
            if sheet[f'{letter}{i}'].value >0:
                trail.tags.add(tags[j])

    return redirect("trell:dashboard")



def change_users(request):
    users = User.objects.all()
    trails = Trail.objects.all()
    trail_count = trails.count()
    user_count = users.count()
    trail_per_user = trail_count//user_count
    print(trail_per_user)
    j = 0
    no = 0
    for trail in trails:
        if no > trail_per_user and j <user_count - 1:
            j+=1
            no = 0 
        user = users[j]
        trail.user = user
        trail.save()
        no+=1
    return redirect("trell:dashboard")
import urllib.request, urllib.parse, urllib.error
from bs4 import BeautifulSoup
import ssl
from urllib.request import Request
import requests

# Ignore SSL certificate errors
def load_images(request):
    tags = Tags.objects.all()
    for tag in tags:
        trails =  Trail.objects.filter(tags = tag, image=None).order_by('created')
        tag = tag.tag.lower()
        tag.replace(" ", '+')
        url = f"https://pixabay.com/api/"
        res = requests.get(url, params={
            "key":"21306606-c6eb388df60ea6f319d5318e0",
            "q":tag, 
            'per_page':100, 
            'image_type':'photo'})
        print({
            "key":"21306606-c6eb388df60ea6f319d5318e0",
            "q":tag, 
            'per_page':100, 
            'image_type':'photo'})
        print(res)
        response = res.json()
        print(response)
        # Retrieve all of the anchor tags
        j = 0
        hits = response['hits']
        for img in hits:
            if j>=trails.count():
                break
            trail = trails[j]
            trail.image = img.get('webformatURL', None)
            print(trail.image)
            trail.save()
            j+=1
    return redirect("trell:dashboard")


def all_tags(request):
    tags = Tags.objects.all()
    for tag in tags:
        cnt= tag.trail_set.all().count()
        print(tag.tag, cnt)